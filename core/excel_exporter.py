"""
Экспорт спецификации в Excel - две вкладки, как в референсном комплекте
конструкторской документации: "Лист" (листовые детали обшивки) и "Трубы"
(профильная труба каркаса). Группировка по секциям (модулям), обозначения
по схеме "К 01.NN.NNN" (проект.модуль.деталь) - упрощённый аналог схемы
"К 01.NN.NN.NNN" из референса (там 4 уровня: проект.секция.узел.деталь,
у нас нет отдельной сущности "узел" внутри модуля, поэтому 3 уровня).

Третья вкладка референса ("Учёт корректировок") не воспроизводится - это
журнал правок конкретного заказа у стороннего производителя, не типовая
структура, которую имеет смысл копировать для наших расчётов.
"""
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import Config
from core.order_paths import order_file

CODE_PREFIX = "К 01"   # префикс обозначения (К - от "Кристалл")

# Стили создаются лениво внутри функций (не на уровне модуля), чтобы импорт
# этого файла не требовал немедленно рабочий openpyxl - важно для тестов/
# песочниц со стаб-заглушками openpyxl.


def _header_style():
    return PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid"), Font(color="FFFFFF", bold=True)


def _section_style():
    return PatternFill(start_color="D9E2EC", end_color="D9E2EC", fill_type="solid"), Font(bold=True, italic=True)


def _border():
    thin = Side(style="thin", color="BBBBBB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _part_note(part):
    """Формирует примечание по вырезам/гибам детали (практическая информация
    для цеха, вместо пустой графы 'корректировка' как в референсе)"""
    notes = []
    if getattr(part, 'cutouts', None):
        labels = ", ".join(c.label or c.shape for c in part.cutouts)
        notes.append(f"вырезы: {labels}")
    if getattr(part, 'bend_lines', None):
        edges = sorted(set(b.edge for b in part.bend_lines))
        notes.append(f"гиб {part.bend_lines[0].offset}мм по краям: {', '.join(edges)}")
    return "; ".join(notes) if notes else ""


def _get_sections(module):
    """
    Возвращает список (название_секции, модуль) для группировки.
    Для KitchenProject - по одной секции на каждый модуль проекта.
    Для одиночного модуля - одна секция с его собственным названием.
    """
    if hasattr(module, 'modules'):  # KitchenProject
        return [(m.name, m) for m in module.modules]
    return [(module.name, module)]


def _write_header(ws, row, headers, col_widths):
    header_fill, header_font = _header_style()
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w


def _write_section_row(ws, row, text, span):
    section_fill, section_font = _section_style()
    border = _border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = section_fill
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = section_fill
        ws.cell(row=row, column=col).border = border


class ExcelExporter:

    @staticmethod
    def export(module):
        Config.init_dirs()
        version = time.strftime("%Y%m%d_%H%M%S")
        filename = order_file(module, "Спецификация.xlsx")

        wb = Workbook()
        ExcelExporter._write_sheet_list(wb, module)
        ExcelExporter._write_sheet_tubes(wb, module)
        wb.active = 0

        wb.save(filename)
        return filename

    @staticmethod
    def _write_sheet_list(wb, module):
        ws = wb.active
        ws.title = "Лист"

        headers = ["Обозначение", "Наименование", "Кол-во", "S, мм", "Примечание"]
        widths = [16, 34, 10, 8, 45]
        _write_header(ws, 1, headers, widths)

        sections = _get_sections(module)
        row = 2
        total_area = 0
        total_qty = 0
        part_num = 0
        border = _border()

        for section_idx, (section_name, mod) in enumerate(sections, 1):
            if not mod.parts:
                continue
            _write_section_row(ws, row, section_name, span=len(headers))
            row += 1

            for part in mod.parts:
                part_num += 1
                code = f"{CODE_PREFIX}.{section_idx:02d}.{part_num:03d}"
                area_total = part.area * part.quantity
                total_area += area_total
                total_qty += part.quantity

                values = [code, part.name, part.quantity, part.thickness, _part_note(part)]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center" if col != 2 and col != 5 else "left",
                                                wrap_text=True, vertical="center")
                row += 1

        # Итоговая строка
        ws.cell(row=row, column=2, value="ИТОГО ЛИСТОВЫХ ДЕТАЛЕЙ").font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_qty).font = Font(bold=True)
        ws.cell(row=row + 1, column=2, value="Общая площадь металла, м²").font = Font(bold=True)
        ws.cell(row=row + 1, column=3, value=round(total_area, 3)).font = Font(bold=True)
        for r in (row, row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).border = border

        ws.freeze_panes = "A2"

    @staticmethod
    def _write_sheet_tubes(wb, module):
        ws = wb.create_sheet("Трубы")

        headers = ["Обозначение", "Наименование", "Кол-во", "Примечание"]
        widths = [16, 34, 10, 30]
        _write_header(ws, 1, headers, widths)

        sections = _get_sections(module)
        row = 2
        total_length_m = 0
        total_qty = 0
        tube_num = 0
        any_tubes = False
        border = _border()

        for section_idx, (section_name, mod) in enumerate(sections, 1):
            tubes = getattr(mod, 'tubes', None) or []
            if not tubes:
                continue
            any_tubes = True
            _write_section_row(ws, row, section_name, span=len(headers))
            row += 1

            for tube in tubes:
                tube_num += 1
                code = f"{CODE_PREFIX}.{section_idx:02d}.Т{tube_num:03d}"
                total_length_m += tube.total_length_mm / 1000
                total_qty += tube.quantity

                values = [code, tube.name, tube.quantity, tube.note]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center" if col != 2 and col != 4 else "left",
                                                wrap_text=True, vertical="center")
                row += 1

        if not any_tubes:
            ws.cell(row=row, column=1, value="(в этом изделии каркас из трубы не используется)")
            row += 1

        ws.cell(row=row, column=2, value="ИТОГО ОТРЕЗКОВ ТРУБЫ").font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_qty).font = Font(bold=True)
        ws.cell(row=row + 1, column=2, value="Общий метраж трубы, м").font = Font(bold=True)
        ws.cell(row=row + 1, column=3, value=round(total_length_m, 2)).font = Font(bold=True)
        for r in (row, row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).border = border

        ws.freeze_panes = "A2"